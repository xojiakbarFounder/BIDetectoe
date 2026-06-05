from __future__ import annotations

import datetime as dt
import threading
from pathlib import Path
from typing import Any

from loguru import logger

from config import settings

_LOCK = threading.Lock()
_HEADERS = [
    "event_id",
    "local_time",
    "object_class",
    "tracker_id",
    "direction",
    "bbox_x1",
    "bbox_y1",
    "bbox_x2",
    "bbox_y2",
]


def _export_path(local_time: dt.datetime) -> Path:
    root = Path(settings.excel_export_dir)
    root.mkdir(parents=True, exist_ok=True)
    return root / f"detection_events_{local_time:%Y-%m-%d}.xlsx"


def append_crossing_event(row: dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        logger.warning("openpyxl is not installed; Excel export skipped.")
        return

    local_time = row["local_time"]
    path = _export_path(local_time)

    with _LOCK:
        if path.exists():
            workbook = load_workbook(path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Detection Events"
            sheet.append(_HEADERS)

        values = []
        for key in _HEADERS:
            value = row.get(key, "")
            if isinstance(value, dt.datetime) and value.tzinfo is not None:
                value = value.replace(tzinfo=None)
            values.append(value)
        sheet.append(values)
        workbook.save(path)
